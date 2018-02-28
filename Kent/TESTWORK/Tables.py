from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.styles.colors import Color
import sqlite3
import parse
import Populate
# conn = sqlite3.connect('/Users/AGuyCalledJP/desktop/databases/final_project/testwork/hoop3.db')
# c = conn.cursor()
# --KEY
# --(FLOR) first Line of rhyme
# --(FPRF) first page illustration found
# --(LPRF) last page illustration found
# --(ILLU) illustrated
# --(ILLT) illustration type
# --(FPIF) first page illustration found
# --(LPIF) last page illustration found
# --(DOPB) date of publication
# --(IllA) illustrator

#Possibly switch contains to illlustrated have that be rID and Illustrator ID

def dad():
	createRhyme()
	createVol()
	createIlli()
	createDrawn()
	createContains()
	createDrawFor()

def createRhyme():
	c = conn.cursor()
	c.execute(''' CREATE TABLE rhyme (
	rID INTEGER,
	flor TEXT NOT NULL,
	PRIMARY KEY (rID)
	);''')

def insertRhyme(flor):
	c.execute('INSERT INTO rhyme VALUES(?,?)', flor,)
	conn.commit()


def createVol():
	c = conn.cursor()
	c.execute(''' CREATE TABLE volume (
	volumeID INTEGER,
	datePublished INTEGER,
	paginated CHECK (paginated = 'Yes' OR paginated = 'No'),
	external TEXT CHECK (external = 'Yes' OR external = 'No'),
	title TEXT NOT NULL,
	PRIMARY KEY (volumeID)
	);''')

def insertVol(values, name):
	l = list(values)
	l.append(name)
	t = tuple(l)
	c.execute('INSERT INTO volume VALUES(?,?,?,?,?)', t,)
	conn.commit()

def createIlli():
	c = conn.cursor()
	c.execute('''CREATE TABLE illustrator (
	illustratorID INTEGER,
	lname TEXT NOT NULL,
	fname TEXT NOT NULL,
	gender TEXT CHECK (gender = 'M' OR gender = 'F' OR gender = 'NULL'),
	dob INTEGER,
	dod INTEGER,
	source1 TEXT,
	source2 TEXT,
	PRIMARY KEY (illustratorID)
	);''')

def insertIlli(values):
	for val in values:
		c.execute('INSERT INTO illustrator VALUES(?,?,?,?,?,?,?,?)', val,)
		conn.commit()


def createDrawn():
	c = conn.cursor()
	c.execute(''' CREATE TABLE drawn (
	rID INTEGER,
	prf INTEGER NOT NULL,
	illt TEXT,
	pif INTEGER NOT NULL,
	volumeID INTEGER NOT NULL,
	FOREIGN KEY (volumeID) REFERENCES volume(volumeID)
	ON UPDATE CASCADE
	ON DELETE CASCADE
	FOREIGN KEY (rID) REFERENCES rhyme(rID)
	ON UPDATE CASCADE
	ON DELETE CASCADE
	);''')

def insertDrawn(values,x):
	hold = (getrID(values[0]))
	total = []
	if values[1] is not 'NULL':
		for v in values[1]:
			if values[3] is not 'NULL':
				for value in values[3]:
					l = list(values)
					l[0] = hold
					l[1] = v
					l[3] = value
					t = tuple(l)
					total.append(t)
			else:
				l = list(values)
				l[0] = hold
				l[1] = v
				t = tuple(l)
				total.append(t)
	for to in total:
		c.execute('INSERT INTO drawn VALUES(?,?,?,?,?rp)', to,)
		conn.commit()


def createContains():
	c = conn.cursor()
	c.execute(''' CREATE TABLE contains (
	volumeID INTEGER NOT NULL,
	rID INTEGER NOT NULL,
	illu TEXT NOT NULL,
	FOREIGN KEY (rID) REFERENCES rhyme(rID)
	ON UPDATE CASCADE
	ON DELETE SET NULL
	FOREIGN KEY (volumeID) REFERENCES volume(volumeID)
	ON UPDATE CASCADE
	ON DELETE SET NULL
	);''')

def insertContains(values):
	val = (values[0], getrID(values[1]),values[2])
	c.execute('INSERT INTO contains VALUES(?,?,?)', val,)
	conn.commit()

def createDrawFor():
	c = conn.cursor()
	c.execute(''' CREATE TABLE drawFor (
	volumeID INTEGER NOT NULL,
	illustratorID INTEGER NOT NULL,
	FOREIGN KEY (volumeID) REFERENCES volume(volumeID)
	ON UPDATE CASCADE
	ON DELETE SET NULL
	FOREIGN KEY (illustratorID) REFERENCES Illustrator(illustratorID)
	ON UPDATE CASCADE
	ON DELETE SET NULL
	);''')

def insertDrawFor(values):
	illustrators = []
	if type(values[1]) is list:
		for value in values[1]:
			hold = getIllId(value)
			tup = (values[0], hold)
			illustrators.append(tup)
	else:
		hold = getIllId(values[1])
		tup = (values[0], hold)
		illustrators.append(tup)
	for illustrator in illustrators:
		c.execute('INSERT INTO drawFor VALUES(?,?)', illustrator,)
		conn.commit()

#"A2","G274"
def fillIllustrator():
	illustrators = parse.readIllu("A2","G274")
	print illustrators

def fillVolume():
	# volumes = parse.volumeInfo()
	x = 1
	Str = "volume "
	parts = parse.readRhyme("A2","G33912")
	for part in parts[3]:
		name = Str + str(x)
		insertVol(part, name)
		x += 1

#"A2","G33912"
def fillRhyme():
	rhymes = parse.readRhyme("A2","G33912")
	rhymes
	# rID = 0
	# flors = []
	# caught = []
	# for rhyme in rhymes[0]:
	# 	if rhyme[0] not in caught:
	# 		caught.append(rhyme[0])
	# 		r = (rID,rhyme[0])
	# 		flors.append(r)
	# 		rID += 1
	# for flor in flors:
	# 	insertRhyme(flor)

def fillContains():
	part = parse.readRhyme("A2","G33912")
	for part in part[2]:
		insertContains(part)

def fillDrawFor():
	part = parse.readRhyme("A2","G33912")
	for part in part[4]:
		insertDrawFor(part)

def fillDrawn():
	part = parse.readRhyme("A2","G33912")
	x = 2
	for part in part[0]:
		insertDrawn(part, x)
		x += 1


def getrID(flor):
	flor = (flor,)
	c.execute('SELECT rID FROM rhyme WHERE flor = ?', flor,)
	ID = c.fetchall()
	return ID[0][0]

def getIllId(ill):
	ill = (ill,)
	c.execute('SELECT illustratorID FROM illustrator WHERE lname = ?', ill)
	ID = c.fetchall()
	if len(ID) == 0:
		return "NULL"
	else:
		return ID[0][0]

def getList():
	part = parse.readRhyme("A2","G33912")
	return part
# def fillDrawn():
# 	global rhymes
# 	for rhyme in rhymes 

# def findVolume(rhyme):
# 	c = conn.cursor()

# def read():
# 	wb = load_workbook(filename = 'hoop.xlsx')
# 	ws = wb["Illustrator info"]
# 	row = ws.max_row 
# 	column = ws.max_column
# 	createIlli()
# 	for x in range(0,row)
# 	print row, column
