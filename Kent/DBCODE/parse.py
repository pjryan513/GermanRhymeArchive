from openpyxl import load_workbook
import sqlite3
import os
import re
import unicodedata
#/Users/AGuyCalledJP/desktop/databases/final_project/testwork/mothergoose
#/Users/AGuyCalledJP/desktop/databases/final_project/testwork/xlsx

#takes a string of the form [x] and returns x
#@Param: String element
def braces(element):
	if '[' in element:
		temp = element.split('[')
		element = temp[1]
	if ']' in element:
		temp = element.split(']')
		element = temp[0]
	return element

def B(prf):
	global paginated
	paginated = 'No'
	prfEnd = []
	if ',' in prf:
		pages = prf.split(',')
		if len(pages) > 1:
			for page in pages:
				if '[' or ']' in page:
					paginated = 'Yes'
					temp = braces(page)
					prfEnd.append(temp)
				else: 
					temp = page
					prfEnd.append(temp)
			return [prfEnd,False]
		else:
			if '[' or ']' in page:
				paginated = 'Yes'
				temp = braces(page)
				prfEnd.append(temp)
			else: 
				temp = page
				prfEnd.append(temp)
			return [prfEnd,False]
	elif '-' in prf:
		pages = prf.split('-')
		if len(pages) > 1:
			for page in pages:
				if '[' or ']' in page:
					paginated = 'Yes'
					temp = braces(page)
					prfEnd.append(temp)
				else: 
					temp = page
					prfEnd.append(temp)
			return [prfEnd,False]
		else:
			if '[' or ']' in page:
				paginated = 'Yes'
				temp = braces(page)
				prfEnd.append(temp)
			else: 
				temp = page
				prfEnd.append(temp)
			return [prfEnd,False]
	elif  re.search('[a-zA-Z]', prf) is None:
		if '[' or ']' in prf:
			pagniate = 'Yes'
			temp = braces(prf)
			prfEnd.append(temp)
		else:
			temp = braces(prf)
			prfEnd.append(temp)
		return [prfEnd,False]
	else:
		return [prf,True]


#function to get the pages the illustration was found on 
#values to be the range of the pages the illustrations were found on
#@Param: String fpoi (first page of illustration)
def E(pif):
	pifEnd = []
	if ',' in pif:
		pages = pif.split(',')
		if len(pages) > 1:
			for page in pages:
				if '[' or ']' in page:
					temp = braces(page)
					pifEnd.append(temp)
				else: 
					temp = page
					pifEnd.append(temp)
			return pifEnd
		else:
			if '[' or ']' in page:
				temp = braces(page)
				pifEnd.append(temp)
			else: 
				temp = page
				pifEnd.append(temp)
			return pifEnd
	elif '-' in pif:
		pages = pif.split('-')
		if len(pages) > 1:
			for page in pages:
				if '[' or ']' in page:
					temp = braces(page)
					pifEnd.append(temp)
				else: 
					temp = page
					pifEnd.append(temp)
			return pifEnd
		else:
			if '[' or ']' in page:
				temp = braces(page)
				pifEnd.append(temp)
			else: 
				temp = page
				pifEnd.append(temp)
			return pifEnd
	else:
		if '[' or ']' in pif:
			temp = braces(pif)
			pifEnd.append(temp)
		else:
			temp = braces(pif)
			pifEnd.append(temp)
		return pifEnd

#Split apart different forms of author data
#@Param: String illustrate
def G(illustrate):
	if "/" in illustrate:
		l = []
		temp = illustrate.split('/')
		for i in range(0,len(temp)):
			l.append(temp[i])
		return l
	elif ";" in illustrate:
		l = []
		temp = illustrate.split(';')
		for i in range(0,len(temp)):
			l.append(temp[i])
		return l
	elif "," in illustrate:
		l = []
		temp = illustrate.split(',')
		for i in range(0,len(temp)):
			l.append(temp[i])
		return l
	elif "and" in illustrate:
		l = []
		temp = illustrate.split('and')
		for i in range(0,len(temp)):
			l.append(temp[i])
		return l
	else:
		return illustrate

#main script function 
def readRhymeInfo(wb, start, end):
	global flor #first line of rhyme
	global fprf #first page rhyme found
	global lprf #last page rhyme found
	global illu #illustrated
	global illt #illustration type
	global fpif #first page illustration found
	global lpif #last page illustration found
	global datePublished 
	global illustrator
	global paginated
	global external
	flip = 0
	lastDate = 0
	volumeID = 0
	trigger = 0
	rhymesIn = []
	illustratedVol = []
	illustratorsOfVol = []
	rhymes = []
	errorReport = []
	ws = wb["rhyme instance"]
	cell_range = ws[start:end] 
	rID = 0
	rowNumber = 2
	for row in cell_range: # This is iterating through rows 1-7
		for cell in row: # This iterates through the columns(cells) in that row
			# print str(cell.value), "|", 
			if cell.column == 'A':
				if '*' in unicode(cell.value):
					rhyme = flor.split("*")
					flor = unicode(rhyme[0])
				else:
					flor = unicode(cell.value) #First line of the rhyme
			if cell.column == 'B':
				if type(cell.value) is unicode:
					fpor = unicodedata.normalize('NFKD', cell.value).encode('ascii','ignore')
				else:
					fpor = str(cell.value)
				hold = B(fpor)
				if hold[1] is True:
					prf = 'NULL'
					if rowNumber not in errorReport:
						errorReport.append(rowNumber)
						continue
				else:
					prf = hold[0]
			if cell.column == 'C':
				if type(cell.value) == str:
					illu = str(cell.value) #Is the rhyme illustrated (y/n)
				else:
					illu = 'NULL'
			if cell.column == 'D':
				if cell.value:
					u = unicode(cell.value)
					hold = unicodedata.normalize('NFKD', u).encode('ascii','ignore')
					if "b/w" in hold:
						illt = unicode(cell.value)
					elif "color" in hold:
						illt = unicode(cell.value)
					elif "color plate" in hold:
						illt = unicode(cell.value)
					else:
						illt = 'NULL'
						if rowNumber not in errorReport:
							errorReport.append(rowNumber)
			if cell.column == 'E':
				if cell.value:
					u = unicode(cell.value)
					hold = unicodedata.normalize('NFKD', u).encode('ascii','ignore')
					if re.search('[a-zA-Z]', hold) == None:
						pif = unicode(cell.value)
					else:
						pif = 'NULL'
						if rowNumber not in errorReport and illu is not 'yes':
							errorReport.append(rowNumber)
							continue
				else:
					pif = "NULL"
			if cell.column == 'F':
				if cell.value:
					date = str(cell.value)
					if flip == 0:
						lastDate = date
						flip += 1
						rhymesIn.append((volumeID, flor, illu))
					elif date != lastDate:
						lastDate = date
						volumeID += 1
						trigger = 0
						rhymesIn.append((volumeID, flor, illu))
					else:
						rhymesIn.append((volumeID, flor, illu))
					if '[' in date:
						external = 'Yes'
						datePublished = braces(date)
					else:
						external = 'No'
						datePublished = str(cell.value) #Get date volume published from external source
			if cell.column == 'G':
				illustrator = unicode(cell.value)
				illustrator = G(illustrator)
				if trigger == 0:
					illustratedVol.append((volumeID, datePublished, paginated))
					illustratorsOfVol.append((volumeID,illustrator))
					trigger = 1
		rhymes.append((flor, illt, volumeID, illu))
		rID += 1
		rowNumber += 1
	return [rhymes, errorReport, rhymesIn, illustratedVol, illustratorsOfVol]

def readRhyme(start, end):
	wb = load_workbook(filename = 'updatedHoop.xlsx')
	rhyme = readRhymeInfo(wb, start, end)
	return rhyme

def removeBrackets(candidate):
	if " " in candidate:
		holder = candidate.split(" ")
		Str = ""
		for s in holder:
			Str += removeBrackets(s)
			Str += " "
		return Str
	elif '[' in candidate:
		Str = ""
		holder = candidate.split('[')
		for i in range(0,len(holder)):
			Str += holder[i]
		return removeBrackets(Str)
	elif ']' in candidate:
		Str = ""
		holder = candidate.split(']')
		for i in range(0,len(holder)):
			Str += holder[i]
		return removeBrackets(Str)
	else:
		return candidate

def parseDOB(value):
	if '?' in value:
		if len(value) is 1:
			dob = "NULL"
			return dob
		else:
			holder = value.split('?')
			dob = int(holder[0])
			return dob
	elif '/' in value:
		holder = value.split('/')
		dob = value[0]
		return dob
	elif "c." in value:
		holder = value.split("c.")
		dob = int(holder[1])
		return dob
	elif '.' in value:
		dod = "NULL"
		return dob
	else:
		dob = int(value)
		return dob

def parseDOD(value):
	if '?' in value:
		if len(value) is 1:
			dod = "NULL"
			return dod
		else:
			holder = value.split('?')
			dod = int(holder[0])
			return dod
	elif '/' in value:
		holder = value.split('/')
		dod = value[0]
		return dod
	elif '.' in value:
		dod = "NULL"
		return dod
	else:
		dod = int(value)
		return dod

def readIllustratorInfo(wb, start, end):
	ws = wb["Illustrators"]
	idCounter = 0
	tableI = []
	cell_range = ws[start:end] 
	for row in cell_range: # This is iterating through rows 1-7
		for cell in row:
			# print str(cell.value), "|", 
			if cell.column == 'A':
				if cell.value:
					x = cell.value
					y = unicode(x)
					lname = removeBrackets(y)
				else:
					lname = "NULL"
			if cell.column == 'B':
				if cell.value:
					x = cell.value
					y = unicode(x)
					fname = removeBrackets(y)
				else:
					fname = "NULL"
			if cell.column == 'C':
				value = str(cell.value)
				if value == '?':
					gender = "NULL"
				else:
					gender = value
			if cell.column == 'D':
				if cell.value:
					value = str(cell.value)
					dob = parseDOB(value)
				else:
					dob = "NULL"
			if cell.column == 'E':
				if cell.value:
					value = str(cell.value)
					dod = parseDOD(value)
				else:
					dod = "NULL"
			if cell.column == 'F':
				if cell.value:
					source1 = str(cell.value)
				else:
					source1 = "NULL"
			if cell.column == 'G':
				if cell.value:
					source2 = str(cell.value)
				else:
					source2 = "NULL"
		tableI.append((idCounter,lname, fname, gender, dob, dod, source1, source2))
		idCounter += 1
	return tableI

def readIllu(start, end):
	wb = load_workbook(filename = 'illustrators.xlsx')
	illustrate = readIllustratorInfo(wb, start, end)
	return illustrate

def removeCurlyBrackets(candidate):
	if " " in candidate:
		holder = candidate.split(" ")
		Str = ""
		for s in holder:
			Str += removeCurlyBrackets(s)
			Str += " "
		return Str
	elif '{' in candidate:
		Str = ""
		holder = candidate.split('{')
		for i in range(0,len(holder)):
			Str += holder[i]
		return removeCurlyBrackets(Str)
	elif '}' in candidate:
		Str = ""
		holder = candidate.split('}')
		for i in range(0,len(holder)):
			Str += holder[i]
		return removeCurlyBrackets(Str)
	else:
		return candidate

def splitAuthor(authors):
	authors = removeCurlyBrackets(authors)
	authors = removeBrackets(authors)
	final = []
	if ',' in authors:
		authors = authors.split(',')
		for author in authors:
			if 'and' in author:
				author = author.split('and')
				for a in author:
					final.append(a)
			else:
				final.append(author)
		final = final[:-1]
		return final
	else:
		return authors


def readVolumeInfo(files):
	volumeID = 0
	address = "NULL"
	title = "NULL"
	language = "NULL"
	publisher = "NULL"
	author = "NULL"
	editor = "NULL"
	year = "NULL"
	note = "NULL"
	# pages = "NULL"
	edition = "NULL"
	series = "NULL"
	volume = "NULL"
	url = "NULL"
	collaborator = "NULL"
	volumes = []
	authors = []
	for file in files:
		for i in file:
			if '=' in i:
				info = i.split('=')
				if info[0] == 'address ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					address = Str
				if info[0] == 'title ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					title = Str
				if info[0] == 'language ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					language = Str
				if info[0] == 'publisher ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					publisher = Str
				if info[0] == 'author ':
					author = splitAuthor(info[1])
					authors.append([volumeID,author])
				if info[0] == 'editor ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					editor = Str
				if info[0] == 'year ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					year = Str
				if info[0] == 'note ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					note = Str
				# if info[0] == 'pages ':
				# 	Str = str(info[1])
				# 	Str = Str[:-2]
				# 	Str = Str[1:]
				# 	Str = removeCurlyBrackets(Str)
				# 	pages = Str
				if info[0] == 'edition ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					edition = Str
				if info[0] == 'series ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					series = Str
				if info[0] == 'volume ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					volume = Str
				if info[0] == 'url ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					url = Str
				if info[0] == 'collaborator ':
					Str = str(info[1])
					Str = Str[:-2]
					Str = Str[1:]
					Str = removeCurlyBrackets(Str)
					collaborator = Str
		volumes.append((volumeID, address, title, language, publisher, editor, year, note, edition, series, volume, url, collaborator))
		volumeID += 1
	return [volumes, authors]


def volumeInfo():
	mothergoose = []
	directory = os.path.normpath("/Users/AGuyCalledJP/desktop/databases/final_project/testwork/mothergoose")
	for subdir, dirs, files in os.walk(directory):
		for file in files:
			lines = []
			if file.endswith(".txt"):
				f=open(os.path.join(subdir, file),'r')
				a = f.readlines()
				for line in a:
					lines.append(line)
				f.close()
				mothergoose.append(lines)
	volumes = readVolumeInfo(mothergoose)
	for volume in volumes:
		print volume, "\n"






